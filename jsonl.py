import csv
import json
import pandas as pd
import openpyxl
import pyarrow.parquet as pq
import pyarrow.orc as orc
import pyarrow.feather as feather
import sqlite3
import xml.etree.ElementTree as ET
import yaml
import msgpack


class jsonl:


    def __init__(self,jsonl_data = []):
        self.data = jsonl_data

    def __str__(self):
        return self.data

    # Override the __len__ method to customize the behavior of len() on this object
    def __len__(self):
        return len(self.data)

    

    def save_jsonl(self, output_file_name):
        with open(output_file_name, 'w') as json_file:
            # Write each JSON object as a separate line
            for json_object in self.data:
                json_file.write(json.dumps(json_object) + '\n')

        print(f"Conversion complete. Output written to {output_file_name}")
        
        return self

    def to_csv(self, output_csv_file):

        with open(output_csv_file, 'w', newline='') as csv_file:
            csv_writer = csv.DictWriter(csv_file, fieldnames=self.data[0].keys())

            # Write the header
            csv_writer.writeheader()

            # Write each JSON object as a row
            csv_writer.writerows(self.data)

        print(f"Conversion complete. Output written to {output_csv_file}")
        return self

    


    def to_xl(self, output_xlsx_file):
        wb = openpyxl.Workbook()
        ws = wb.active

        # Write the header
        header = list(self.data[0].keys())
        for col_num, value in enumerate(header, 1):
            ws.cell(row=1, column=col_num, value=value)

        # Write each JSON object as a row
        for row_num, obj in enumerate(self.data, 2):
            for col_num, value in enumerate(obj.values(), 1):
                ws.cell(row=row_num, column=col_num, value=value)

        wb.save(output_xlsx_file)
        print(f"Conversion to Excel complete. Output written to {output_xlsx_file}")
        return self

    def to_dataframe(self):
        return pd.DataFrame(self.data)


    def to_orc(self, output_orc_file):
        df = pd.DataFrame(self.data)
        table = orc.Table.from_pandas(df)
        orc.write_table(table, output_orc_file)
        print(f"Conversion to ORC complete. Output written to {output_orc_file}")
        return self

    def to_hdf5(self, output_hdf5_file):
        df = pd.DataFrame(self.data)
        df.to_hdf(output_hdf5_file, key="data", format="table", mode="w")
        print(f"Conversion to HDF5 complete. Output written to {output_hdf5_file}")
        return self

    def to_parquet(self, output_parquet_file):
        df = pd.DataFrame(self.data)
        df.to_parquet(output_parquet_file, index=False)
        print(f"Conversion to Parquet complete. Output written to {output_parquet_file}")
        return self

    def to_avro(self, output_avro_file):
        df = pd.DataFrame(self.data)
        df.to_avro(output_avro_file, index=False)
        print(f"Conversion to Avro complete. Output written to {output_avro_file}")
        return self

    def to_xml(self, output_xml_file):
        root = ET.Element("root")
        for item in self.data:
            element = ET.SubElement(root, "item")
            for key, value in item.items():
                child = ET.SubElement(element, key)
                child.text = str(value)
        tree = ET.ElementTree(root)
        tree.write(output_xml_file, xml_declaration=True, encoding='utf-8', method="xml")
        print(f"Conversion to XML complete. Output written to {output_xml_file}")
        return self

    def to_yaml(self, output_yaml_file):
        with open(output_yaml_file, 'w') as yaml_file:
            yaml.dump(self.data, yaml_file, default_flow_style=False)
        print(f"Conversion to YAML complete. Output written to {output_yaml_file}")
        return self

    def to_json(self, output_json_file):
        with open(output_json_file, 'w') as json_file:
            json.dump(self.data, json_file, indent=2)
        print(f"Conversion to JSON complete. Output written to {output_json_file}")
        return self

    def to_feather(self, output_feather_file):
        df = pd.DataFrame(self.data)
        df.to_feather(output_feather_file)
        print(f"Conversion to Feather complete. Output written to {output_feather_file}")
        return self

    def to_sqlite(self, output_sqlite_file, table_name='data'):
        df = pd.DataFrame(self.data)
        conn = sqlite3.connect(output_sqlite_file)
        df.to_sql(table_name, conn, index=False, if_exists="replace")
        print(f"Conversion to SQLite complete. Output written to {output_sqlite_file}")
        return self

    def to_msgpack(self, output_msgpack_file):
        with open(output_msgpack_file, "wb") as f:
            f.write(msgpack.packb(self.data))
        print(f"Conversion to Msgpack complete. Output written to {output_msgpack_file}")
        return self


    

    

def csv_to_jsonl( input_csv_file):


        jsl = jsonl()
        with open(input_csv_file, 'r') as csv_file:
            csv_reader = csv.DictReader(csv_file)

            # Get the fieldnames (keys) from the header row of the CSV
            fieldnames = csv_reader.fieldnames
            
            # Iterate through each row in the CSV file
            for row in csv_reader:
                # Create a dictionary with the keys and values
                json_object = {field: row[field] for field in fieldnames}

                # Append the dictionary to the list
                jsl.data.append(json_object)

        return jsl


def read_jsonl(input_file_name):

        jsl = jsonl()
        with open(input_file_name, 'r') as json_file:
            # Read each line from the file
            json_lines = json_file.readlines()

            # Convert each line to a dictionary
            jsl.data = [json.loads(line.strip()) for line in json_lines]

        return jsl


def dataframe_to_jsonl(df):
        result_list = df.to_dict(orient='records')
        jsl = jsonl(result_list)
        return jsl


def xl_to_jsonl(self, input_excel_file, sheet_name=0):
        df = pd.read_excel(input_excel_file, sheet_name=sheet_name)
        
        return dataframe_to_jsonl(df)

def avro_to_jsonl(input_avro_file):
    df = pd.read_avro(input_avro_file)
    return dataframe_to_jsonl(df)

def parquet_to_jsonl(input_parquet_file):
    df = pd.read_parquet(input_parquet_file)
    return dataframe_to_jsonl(df)

def orc_to_jsonl(input_orc_file):
    table = orc.read_table(input_orc_file)
    return dataframe_to_jsonl(table.to_pandas())

def feather_to_jsonl(input_feather_file):
    table = feather.read_table(input_feather_file)
    return dataframe_to_jsonl(table.to_pandas())

def xml_to_jsonl(input_xml_file):
    tree = ET.parse(input_xml_file)
    root = tree.getroot()
    result_list = []

    for item in root.findall('item'):
        data = {child.tag: child.text for child in item}
        result_list.append(data)

    jsl = jsonl(result_list)
    return jsl

def yaml_to_jsonl(input_yaml_file):
    with open(input_yaml_file, 'r') as yaml_file:
        data = yaml.safe_load(yaml_file)
    jsl = jsonl(data)
    return jsl

def json_to_jsonl(input_json_file):
    with open(input_json_file, 'r') as json_file:
        data = json.load(json_file)
    jsl = jsonl(data)
    return jsl


def from_msgpack_to_dict_list(input_msgpack_file):
    with open(input_msgpack_file, 'rb') as msgpack_file:
        data = msgpack.unpack(msgpack_file, encoding='utf-8')
    jsl = jsonl(data)
    return jsl

def sqlite_to_jsonl(input_sqlite_file, table_name='data'):
    conn = sqlite3.connect(input_sqlite_file)
    query = f"SELECT * FROM {table_name};"
    df = pd.read_sql_query(query, conn)
    return dataframe_to_jsonl(df)
  
def hdf5_to_jsonl(input_hdf5_file, key='data'):
    df = pd.read_hdf(input_hdf5_file, key=key)
    return dataframe_to_jsonl(df)

